from openpyxl import load_workbook
#from .models import *

workbook = load_workbook(filename="../data/laporan_lb1_01112019_101607.xls", read_only=True)
sheet = workbook.active

x=[]

for row in sheet.iter_rows(min_row=15, values_only=True):
	x.append(row[2])

print(x)
